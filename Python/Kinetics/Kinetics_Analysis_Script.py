########################################################################################################################
# Created on Tue 2024_02_20
# Created by Melissa Franke
# Last edit: 2026_03_31
########################################################################################################################

import csv
from math import ceil, floor
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from os import scandir
from pybaselines.polynomial import imodpoly
from PyQt6.QtWidgets import QFileDialog
from scipy import signal, integrate
from scipy.optimize import curve_fit

# # # # # # # # # # # # # # # # # #
# # # # #    functions    # # # # #
# # # # # # # # # # # # # # # # # #

def parse_data(reader):
    error_occurred=0
    empty_column=0
    ROI_list=[]
    raw_data=[]
    time=[]
    if reader[0][-1]=='' or reader[0][-1]==None:
        empty_column=1
    for j in range(0,len(reader[0])-empty_column):
        column=[]
        for k in range(0,len(reader)):
            x=reader[k][j]
            if x==None:
                break
            elif x=='':
                x=np.nan
            else:
                try:
                    x=float(x)
                except:
                    x=x
            column.append(x)
        if len(column)==0:
            error_occurred=1
        if not error_occurred:
            while np.isnan(column[-1]): #nan-pruning
                column.pop()
            ROI_list.append(column[0])
            column.pop(0)
            raw_data.append(column)
            time.append(list([i for i in range(1,len(column)+1)]))
    return raw_data, ROI_list, time

def calculate_framerate_from_column(time_column):
    diff=time_column[1]-time_column[0]
    return 1/diff

def correct_baseline(signal):
    baseline=imodpoly(signal,poly_order=2,use_original=True)
    baseline=baseline[0]
    corrected=list([signal[i]-baseline[i] for i in range(0,len(signal))])
    return corrected

def find_peaks_threshold(data,threshold):
    peak_ind=list()
    peak_val=list()
    mask=[i-threshold if i>threshold else 0 for i in data]
    peak_area=find_areas_above_threshold(mask)
    clustered_peaks=cluster_peaks(peak_area)
    clustered_peaks=prune_clusters(clustered_peaks,data)
    peak_ind=find_peaks_in_cluster(clustered_peaks,mask)
    for index in peak_ind:
        peak_val.append(data[index])
    return peak_ind, peak_val
    #return index, value

def find_points(data,peak_inds,threshold):
    error=0
    indices=list()
    values=list()
    temp_begin=list()
    temp_end=list()
    x=[i for i in range(0,len(data))]
    crossings=cstm_intersection(x,data,threshold)   #get all approximate indices crossing data
    if len(crossings)!=0:                           #choose correct crossing according to data and threshold relation
        for index in crossings:
            if data[floor(index)]<data[ceil(index)]:
                indices.append(floor(index))
            else:
                indices.append(ceil(index))
        
        indices=list(set(indices))
        indices.sort()
        pop_adapter=0
        for i in range(0,len(indices)): #get rid of all crossings that are peak 
            if indices[i-pop_adapter] in peak_inds:
                indices.pop(i-pop_adapter)
                pop_adapter+=1

        #check if every peak has min 1 point before and after
        indices=indices+peak_inds
        indices.sort()
        problem=False
        if indices[0] in peak_inds:
            problem=True
        elif indices[len(indices)-1] in peak_inds:
            problem=True
        else:
            for i in range(1,len(indices)):
                if indices[i] in peak_inds and indices[i-1] in peak_inds:
                    problem=True
                    break

        if not problem: #every peak has at least one beginning and one end
            for i in range(0,len(indices)):
                if indices[i] in peak_inds:
                    temp_begin.append(indices[i-1])
                    temp_end.append(indices[i+1])
        else:   #calculate approximate, then let user choose to keep or not
            for i in range(0,len(peak_inds)):
                if i==0 and len(peak_inds)==1:
                    rise_start=0
                    rise_end=peak_inds[i]
                    fall_start=peak_inds[i]+1
                    fall_end=x[-1]
                elif i==0 and len(peak_inds)!=1:
                    rise_start=0
                    rise_end=peak_inds[i]
                    fall_start=peak_inds[i]+1
                    fall_end=peak_inds[i+1]
                elif i==len(peak_inds)-1:
                    rise_start=peak_inds[i-1]+1
                    rise_end=peak_inds[i]
                    fall_start=peak_inds[i]+1
                    fall_end=x[-1]
                else:
                    rise_start=peak_inds[i-1]+1
                    rise_end=peak_inds[i]
                    fall_start=peak_inds[i]+1
                    fall_end=peak_inds[i+1]
                    
                if rise_start==rise_end and fall_start==fall_end:
                    temp_begin.append(rise_start)
                    temp_end.append(fall_start)

                elif rise_start==rise_end:
                    temp_begin.append(rise_start)
                    fall_max=max(data[fall_start:fall_end])
                    fall_min=min(data[fall_start:fall_end])
                    if fall_min<=threshold<=fall_max: #crossing exists in indices
                        temp_end.append(indices[indices.index(peak_inds[i])+1])
                    elif threshold<=fall_min:    #crossing has to be approximated
                        temp_end.append(peak_inds[i]+data[fall_start:fall_end].index(fall_min))
                    elif threshold>=fall_max:    #crossing has to be approximated
                        temp_end.append(peak_inds[i]+data[fall_start:fall_end].index(fall_max))

                elif fall_start==fall_end:
                    temp_end.append(fall_start)
                    rise_max=max(data[rise_start:rise_end])
                    rise_min=min(data[rise_start:rise_end])
                    if rise_min<=threshold<=rise_max: #crossing exists in indices
                        temp_begin.append(indices[indices.index(peak_inds[i])-1])
                    elif threshold<=rise_min:    #crossing has to be approximated
                        data_snippet=data[rise_start:rise_end]
                        data_snippet.reverse()
                        index=len(data_snippet)-1-data_snippet.index(rise_min)
                        temp_begin.append(rise_start+index)
                    elif threshold>=rise_max:    #crossing has to be approximated
                        data_snippet=data[rise_start:rise_end]
                        data_snippet.reverse()
                        index=len(data_snippet)-1-data_snippet.index(rise_max)
                        temp_begin.append(rise_start+index)
                else:                    
                    rise_max=max(data[rise_start:rise_end])
                    rise_min=min(data[rise_start:rise_end])
                    fall_max=max(data[fall_start:fall_end])
                    fall_min=min(data[fall_start:fall_end])
                    if rise_min<=threshold<=rise_max: #crossing exists in indices
                        temp_begin.append(indices[indices.index(peak_inds[i])-1])
                    elif threshold<=rise_min:    #crossing has to be approximated
                        data_snippet=data[rise_start:rise_end]
                        data_snippet.reverse()
                        index=len(data_snippet)-1-data_snippet.index(rise_min)
                        temp_begin.append(rise_start+index)
                    elif threshold>=rise_max:    #crossing has to be approximated
                        data_snippet=data[rise_start:rise_end]
                        data_snippet.reverse()
                        index=len(data_snippet)-1-data_snippet.index(rise_max)
                        temp_begin.append(rise_start+index)
                    if fall_min<=threshold<=fall_max: #crossing exists in indices
                        temp_end.append(indices[indices.index(peak_inds[i])+1])
                    elif threshold<=fall_min:    #crossing has to be approximated
                        temp_end.append(peak_inds[i]+data[fall_start:fall_end].index(fall_min))
                    elif threshold>=fall_max:    #crossing has to be approximated
                        temp_end.append(peak_inds[i]+data[fall_start:fall_end].index(fall_max))
                
        values_begin=[data[j] for j in temp_begin]
        values_end=[data[j] for j in temp_end]
    return temp_begin,values_begin,temp_end,values_end

def calc_fwhm(data,begin,peak_val,end,amplitude,framerate):
    x1=list()
    y1=list()
    hm=peak_val-(amplitude/2)
    for j in range(begin,end+1):
        x1.append(j)
        y1.append(data[j])
    x_border=cstm_intersection(x1,y1,hm)
    if len(x_border)<2:
        width=None
    else:
        border_left=min(x_border)
        border_right=max(x_border)
        width=(border_right-border_left)/framerate
    return width
    #return fwhm

def area_under_curve(data_snippet,time_snippet):
    return integrate.simpson(y=data_snippet,x=time_snippet)

def calc_risetime8020(data,amplitude,peak_begin_index,peak_begin_value,peak_index, framerate):
    vrisetime=list()
    vx_twenty=list()
    vy_twenty=list()
    vx_eighty=list()
    vy_eighty=list()
    for i in range(0,len(peak_index)):
        value1=peak_begin_value[i]+0.2*amplitude[i]
        value2=peak_begin_value[i]+0.8*amplitude[i]
        c=0
        x=list()
        y_data=list()
        for j in data[peak_begin_index[i]:peak_index[i]+1]:
            x.append(c)
            y_data.append(j)
            c+=1
        x_twenty=cstm_intersection(x,y_data,value1)
        if len(x_twenty)!=0:
            x_twenty=peak_index[i]+min(x_twenty)
            y_twenty=value1
        else:
            x_twenty=np.nan
            y_twenty=np.nan
        x_eighty=cstm_intersection(x,y_data,value2)
        if len(x_eighty)!=0:
            x_eighty=peak_index[i]+min(x_eighty)
            y_eighty=value2
        else:
            x_eighty=np.nan
            y_eighty=np.nan

        if not np.isnan(x_eighty):
            vrisetime.append((x_eighty-x_twenty)/framerate)
        else:
            vrisetime.append(np.nan)
        vx_twenty.append(x_twenty)
        vy_twenty.append(y_twenty)
        vx_eighty.append(x_eighty)
        vy_eighty.append(y_eighty)
    
    return vx_twenty, vy_twenty, vx_eighty, vy_eighty, vrisetime

def calc_decay8020(data,amplitude,peak_end_index,peak_end_value,peak_index, framerate):
    decay=list()
    vx_twenty=list()
    vy_twenty=list()
    vx_eighty=list()
    vy_eighty=list()
    for i in range(0,len(peak_index)):
        value1=peak_end_value[i]+0.2*amplitude[i]
        value2=peak_end_value[i]+0.8*amplitude[i]
        c=0
        x=list()
        y_data=list()
        for j in data[peak_index[i]:peak_end_index[i]+1]:
            x.append(c)
            y_data.append(j)
            c+=1
        x_twenty=cstm_intersection(x,y_data,value1)
        if len(x_twenty)!=0:
            x_twenty=peak_index[i]+min(x_twenty)
            y_twenty=value1
        else:
            x_twenty=np.nan
            y_twenty=np.nan
        x_eighty=cstm_intersection(x,y_data,value2)
        if len(x_eighty)!=0:
            x_eighty=peak_index[i]+min(x_eighty)
            y_eighty=value2
        else:
            x_eighty=np.nan
            y_eighty=np.nan
        if not np.isnan(x_eighty):
            decay.append((x_twenty-x_eighty)/framerate)
        else:
            decay.append(np.nan)
        vx_twenty.append(x_twenty)
        vy_twenty.append(y_twenty)
        vx_eighty.append(x_eighty)
        vy_eighty.append(y_eighty)
    return vx_twenty, vy_twenty, vx_eighty, vy_eighty, decay

def calc_tau_percent(decay_area,amplitude,time_data):
    tau=None
    value=decay_area[0]-amplitude*0.632 #percent amplitude
    index=cstm_intersection(time_data,decay_area,value)
    if len(index)!=0:
        index=min(index)
        tau=index-time_data[0]
    return tau

def calc_tau_fit(decay_area,framerate):
    error=False
    tau=None
    time=[i/framerate for i in range(1,len(decay_area))]
    time.insert(0,0)
    try:
        parameters,para_cov=curve_fit(exponential,time,decay_area)
    except RuntimeError:
        error=True
    except:
        error=True
    if not error:
        tau=parameters[1]
    return tau

def calc_inter_spike_interval(peak_time):
    time_between_spikes=[]
    for i in range(1,len(peak_time)):
        time_between_spikes.append(peak_time[i]-peak_time[i-1])
    isi_mean=np.mean(time_between_spikes)
    isi_median=np.median(time_between_spikes)
    isi_error=np.std(time_between_spikes)
    return time_between_spikes,isi_mean,isi_median,isi_error

def calc_frequency(time_between_peaks):
    frequency_mean=1/np.mean(time_between_peaks)
    frequency_median=1/np.median(time_between_peaks)
    frequency_error=1/np.std(time_between_peaks)
    return frequency_mean, frequency_median, frequency_error

# # # # # # # # # # # # # # # # # # # # # 
# # # # #    helper functions   # # # # #
# # # # # # # # # # # # # # # # # # # # #

def find_areas_above_threshold(mask):
    peak_areas=list()       #all above threshold, index
    for i in range(0,len(mask)):
        if mask[i]!=0:
            peak_areas.append(i)
    return peak_areas

def cluster_peaks(peak_areas):
    diff=list()             #differences,time distance
    interval=list()         #all above threshold where time distance=1, index
    clustered_peaks=list()  #list of intervals, index

    diff.append(peak_areas[0])
    for i in range(1,len(peak_areas)):
        diff.append(peak_areas[i]-peak_areas[i-1])

    for i in range(0,len(diff)):    #later merge with diff creation loop: if peak-peak>1 clusterCounter+=1
        if diff[i]==1:
            interval.append(peak_areas[i])
        else:
            if interval!=[]:
                clustered_peaks.append(interval.copy())
            interval.clear()
            interval.append(peak_areas[i])
        if i==len(diff)-1 and interval!=[]:
            clustered_peaks.append(interval.copy())
    return clustered_peaks

def prune_clusters(clustered_peaks,data):
    # pruning of clustering
    if clustered_peaks[0][0]==0:  #first peak at begin recording
        clustered_peaks.pop(0)
    if clustered_peaks!=[]and clustered_peaks[-1][-1]==len(data)-1:  #last peak at end recording
        clustered_peaks.pop(len(clustered_peaks)-1)
    return clustered_peaks

def find_peaks_in_cluster(clustered_peaks,data):
    peak_ind=list()   
    # get peak indices from cluster
    for i in range(0,len(clustered_peaks)):
        peak_ind_temp=clustered_peaks[i][0]
        peak_val_temp=data[clustered_peaks[i][0]]

        for j in range(0,len(clustered_peaks[i])):
            if data[clustered_peaks[i][j]]>peak_val_temp:
                peak_val_temp=data[clustered_peaks[i][j]]
                peak_ind_temp=clustered_peaks[i][j]
        peak_ind.append(peak_ind_temp)
    return peak_ind

def create_derivative(data):
    derivative=list()
    derivative.append(0)
    for i in range(1,len(data)):
        derivative.append(data[i]-data[i-1])
    return derivative

def cstm_intersection(xdata, ydata,value):
    point_inds=list()
    for i in range(1,len(ydata)):
        if (ydata[i-1]<=value and ydata[i]>=value) or (ydata[i-1]>=value and ydata[i]<=value): #crossing threshold
            point_inds.append(xdata[i-1]+((xdata[i]-xdata[i-1])/2))
    return point_inds

def exponential(x,a,t,b):
    return a*np.exp(-x/t) +b

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 

# # # # # # # # # # # # # #
# # # # #    IO   # # # # #
# # # # # # # # # # # # # #

#user input output folder
output_folder=input('Enter path to output directory: ')
if output_folder[-1]=='\\' or output_folder[-1]=='/':
    output_folder.pop()

#user input input folder
filelist=list()
input_folder=input('Enter path to input directory: ')
if input_folder[-1]=='\\' or input_folder[-1]=='/':
    input_folder.pop()

with scandir(input_folder) as dir:
    for element in dir:
        if element.is_file():
            filelist.append(input_folder+'/'+element.name)

#set up visualisation
fig,ax=plt.subplots()

#for all datasets
for i in range(0,len(filelist)):
    framerate=1
    ROI_list=list()
    time=list()
    raw_data=list()
    working_data=list()
    framerate=list()
    #housekeeping
    deleted_columns=list()
    deleted_peaks=list()
    #peaks
    peak_time=list()
    peak_index=list()
    peak_value=list()
    peak_begin_time=list()
    peak_begin_index=list()
    peak_begin_value=list()
    peak_end_time=list()
    peak_end_index=list()
    peak_end_value=list()
    #parameters
    amplitude=list()
    fwhm=list()
    full_width=list()
    auc=list()
    risetime=list()
    decay=list()
    rise20_time=list()
    rise20_index=list()
    rise20_value=list()
    rise80_time=list()
    rise80_index=list()
    rise80_value=list()
    risetime2080=list()
    fall20_time=list()
    fall20_index=list()
    fall20_value=list()
    fall80_time=list()
    fall80_index=list()
    fall80_value=list()
    decay2080=list()
    tau_approx=list()
    tau_fit=list()
    responsetime=list()
    time_between_peaks=list()
    mean_isi=list()
    median_isi=list()
    sd_isi=list()
    mean_freq=list()
    median_freq=list()
    sd_freq=list()
    periodicity=list()
    mean=list()
    sd=list()


    # # # # # # # # # # # # # # # # # #
    # # # # #    parse data   # # # # #
    # # # # # # # # # # # # # # # # # #

    with open(filelist[i]) as file:
        #read in data
        reader=list(csv.reader(file,delimiter=','))
        #parse data
        raw_data, ROI_list, time = parse_data(reader)
        name=filelist[i][filelist[i].rindex('/')+1:-4]
        
        #check first column time
        framerate_calculated=False
        diff=[raw_data[0][i]-raw_data[0][i-1] for i in range(1,len(raw_data[0]))]
        diff_tolerance=np.mean(diff)*0.01
        if all(diff[i-1]-diff[i]<diff_tolerance for i in range(1,len(diff))):
            framerate_temp=calculate_framerate_from_column(raw_data[0])
            framerate_calculated=True
        if framerate_calculated:
            keep=input(f'A framerate has been calculated from the first data column: {framerate_temp} Hz. Keep framerate? [y/n] ')
            if keep=='n':
                framerate_temp=input('Enter new framerate in Hz: ')
                framerate=float(framerate_temp)
                time=list([i/framerate for i in range(1,len(raw_data[0])+1)])
            else:
                framerate=framerate_temp
                time=raw_data[0]
            raw_data.pop(0)
            ROI_list.pop(0)
            del framerate_temp
    #prepared for current dataset: raw data, time, ROI_list, framerate
    column_adapter=0
    for j in range(0,len(ROI_list)):

        # # # # # # # # # # # # # # # # # # # #
        # # # # #    analysis start   # # # # #
        # # # # # # # # # # # # # # # # # # # #
        print(f'dataset: {name} ROI: {ROI_list[j]}')
        #display raw_data
        plt.clf()
        plt.plot(time,raw_data[j-column_adapter])
        plt.show(block=False)
        plt.pause(0.001)

        keep=input('Keep this column? [y,n] ')
        if keep=='n':
            deleted_columns.append(j)
            ROI_list.pop(j-column_adapter)
            raw_data.pop(j-column_adapter)
            column_adapter+=1
        else:
            # # # # # # # # # # # # # # # # # # # #
            # # # # #    container prep   # # # # #
            # # # # # # # # # # # # # # # # # # # #

            working_data.append([])
            deleted_peaks.append([])
            #peaks
            peak_time.append([])
            peak_index.append([])
            peak_value.append([])
            peak_begin_time.append([])
            peak_begin_index.append([])
            peak_begin_value.append([])
            peak_end_time.append([])
            peak_end_index.append([])
            peak_end_value.append([])
            #parameters
            amplitude.append([])
            fwhm.append([])
            full_width.append([])
            auc.append([])
            risetime.append([])
            decay.append([])
            rise20_time.append([])
            rise20_index.append([])
            rise20_value.append([])
            rise80_time.append([])
            rise80_index.append([])
            rise80_value.append([])
            risetime2080.append([])
            fall20_time.append([])
            fall20_index.append([])
            fall20_value.append([])
            fall80_time.append([])
            fall80_index.append([])
            fall80_value.append([])
            decay2080.append([])
            tau_approx.append([])
            tau_fit.append([])
            responsetime.append([])
            time_between_peaks.append([])
            mean_isi.append([])
            median_isi.append([])
            sd_isi.append([])
            mean_freq.append([])
            median_freq.append([])
            sd_freq.append([])
            periodicity.append([])
            mean.append([])
            sd.append([])

            #baseline shift correction
            working_data[j-column_adapter]=correct_baseline(raw_data[j-column_adapter])

            plt.plot(time,working_data[j-column_adapter])
            plt.show(block=False)
            plt.pause(0.001)

            #define baseline range for peak detect
            interval_user=input('Please enter the value range of your baseline as shown <lower_value,higher_value> ')
            interval_user=interval_user.split(',')
            baseline_interval=[float(interval_user[0]),float(interval_user[1])]
            del interval_user
            baseline_slice=[i if baseline_interval[0]<=i<=baseline_interval[1] else np.nan for i in working_data[j-column_adapter]]
            mean[j-column_adapter]=np.nanmean(baseline_slice)
            sd[j-column_adapter]=np.nanstd(baseline_slice)
            del baseline_slice, baseline_interval

            #detect peaks above mean+3sd
            peak_index[j-column_adapter],peak_value[j-column_adapter]=find_peaks_threshold(working_data[j-column_adapter],mean[j-column_adapter]+(3*sd[j-column_adapter]))
            peak_time[j-column_adapter]=list([index/framerate for index in peak_index[j-column_adapter]])

            plt.scatter(peak_time[j-column_adapter],peak_value[j-column_adapter])
            plt.show(block=False)
            plt.pause(0.001)

            peak_ok=input('Do you want to keep these peaks? [y,n] ')
            if peak_ok=='n':
                escape=False
                while not escape:
                    del_list=input('Enter "stop" to continue, Enter deletion range to keep deleting: <x1,x2,y1,y2> ')
                    if del_list=='stop':
                        escape=True
                    else:
                        del_list=del_list.split(',')
                        del_list=list([float(element) for element in del_list])
                        points_found=False
                        pop_adapter=0
                        for k in range(0,len(peak_index[j-column_adapter])):
                            if del_list[0] <= peak_time[j-column_adapter][k-pop_adapter] <= del_list[1] or del_list[0] >= peak_time[j-column_adapter][k-pop_adapter] >= del_list[1]:
                                points_found=True
                                if del_list[2] <= peak_value[j-column_adapter][k-pop_adapter] <= del_list[3] or del_list[2] >= peak_value[j-column_adapter][k-pop_adapter] >= del_list[3]:    
                                    peak_time[j-column_adapter].pop(k-pop_adapter)
                                    peak_index[j-column_adapter].pop(k-pop_adapter)
                                    peak_value[j-column_adapter].pop(k-pop_adapter)
                                    pop_adapter+=1
                                    deleted_peaks[j-column_adapter].append(k)
                                    plt.clf()
                                    plt.plot(raw_data[j-column_adapter])
                                    plt.plot(working_data[j-column_adapter])
                                    plt.scatter(peak_time[j-column_adapter],peak_value[j-column_adapter])
                                    plt.show(block=False)
                                    plt.pause(0.001)
                            else:
                                if points_found:
                                    break

            #detect peak begin and end, first below mean
            peak_begin_index[j-column_adapter],peak_begin_value[j-column_adapter],peak_end_index[j-column_adapter],peak_end_value[j-column_adapter]=find_points(working_data[j-column_adapter],peak_index[j-column_adapter],mean[j-column_adapter])
            peak_begin_time[j-column_adapter]=list([index/framerate for index in peak_begin_index[j-column_adapter]])
            peak_end_time[j-column_adapter]=list([index/framerate for index in peak_end_index[j-column_adapter]])

            plt.scatter(peak_begin_time[j-column_adapter],peak_begin_value[j-column_adapter])
            plt.scatter(peak_end_time[j-column_adapter],peak_end_value[j-column_adapter])
            plt.show(block=False)
            plt.pause(0.001)

            #parameter calculation
            amplitude[j-column_adapter]=list([peak_value[j-column_adapter][k]-(peak_begin_value[j-column_adapter][k]+peak_end_value[j-column_adapter][k])/2 for k in range(0,len(peak_index[j-column_adapter]))])
            fwhm[j-column_adapter]=list([calc_fwhm(working_data[j-column_adapter],peak_begin_index[j-column_adapter][k],peak_value[j-column_adapter][k],peak_end_index[j-column_adapter][k],amplitude[j-column_adapter][k],framerate) for k in range(0,len(peak_index[j-column_adapter]))])
            full_width[j-column_adapter]=list([peak_end_time[j-column_adapter][k]-peak_begin_time[j-column_adapter][k] for k in range(0,len(peak_index[j-column_adapter]))])
            auc[j-column_adapter]=list([area_under_curve(working_data[j-column_adapter][peak_begin_index[j-column_adapter][k]:peak_end_index[j-column_adapter][k]],time[peak_begin_index[j-column_adapter][k]:peak_end_index[j-column_adapter][k]]) for k in range(0,len(peak_index[j-column_adapter]))])
            
            risetime[j-column_adapter]=list([peak_time[j-column_adapter][k]-peak_begin_time[j-column_adapter][k] for k in range(0,len(peak_index[j-column_adapter]))])
            decay[j-column_adapter]=list([peak_end_time[j-column_adapter][k]-peak_time[j-column_adapter][k] for k in range(0,len(peak_index[j-column_adapter]))])
            rise20_index[j-column_adapter],rise20_value[j-column_adapter],rise80_index[j-column_adapter],rise80_value[j-column_adapter],risetime2080[j-column_adapter]=calc_risetime8020(working_data[j-column_adapter],amplitude[j-column_adapter],peak_begin_index[j-column_adapter],peak_begin_value[j-column_adapter],peak_index[j-column_adapter], framerate)
            fall20_index[j-column_adapter],fall20_value[j-column_adapter],fall80_index[j-column_adapter],fall80_value[j-column_adapter],decay2080[j-column_adapter]=calc_decay8020(working_data[j-column_adapter],amplitude[j-column_adapter],peak_end_index[j-column_adapter],peak_end_value[j-column_adapter],peak_index[j-column_adapter], framerate)
                            
            rise20_time[j-column_adapter]=list([element/framerate for element in rise20_index[j-column_adapter]])
            rise80_time[j-column_adapter]=list([element/framerate for element in rise80_index[j-column_adapter]])
            fall20_time[j-column_adapter]=list([element/framerate for element in fall20_index[j-column_adapter]])
            fall80_time[j-column_adapter]=list([element/framerate for element in fall80_index[j-column_adapter]])
            
            tau_approx[j-column_adapter]=list([calc_tau_percent(working_data[j-column_adapter][peak_index[j-column_adapter][k]:peak_end_index[j-column_adapter][k]] ,amplitude[j-column_adapter][k],time[peak_index[j-column_adapter][k]:peak_end_index[j-column_adapter][k]]) for k in range(0,len(peak_index[j-column_adapter]))])
            tau_fit[j-column_adapter]=list([calc_tau_fit(working_data[j-column_adapter][peak_index[j-column_adapter][k]:peak_end_index[j-column_adapter][k]], framerate) for k in range(0,len(peak_index[j-column_adapter]))])    
                        
            time_between_peaks[j-column_adapter],mean_isi[j-column_adapter],median_isi[j-column_adapter],sd_isi[j-column_adapter]=calc_inter_spike_interval(peak_time[j-column_adapter])  
            mean_freq[j-column_adapter],median_freq[j-column_adapter],sd_freq[j-column_adapter]=calc_frequency(time_between_peaks[j-column_adapter])

            if sd_isi[j-column_adapter]==0:
                periodicity[j-column_adapter]=0
            else:
                periodicity[j-column_adapter]=sd_isi[j-column_adapter]/mean_isi[j-column_adapter]

            input('Press Enter for next column.')
    
    # # # # # # # # # # # # # # # # # # # # # # # #
    # # # # #    export all parameters    # # # # #
    # # # # # # # # # # # # # # # # # # # # # # # #

    wb=Workbook()
    ws=wb.active
    abc=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ','CA','CB','CC','CD','CE','CF','CG','CH','CI','CJ','CK','CL','CM','CN','CO','CP','CQ','CR','CS','CT','CU','CV','CW','CX','CY','CZ','DA','DB','DC','DD','DE','DF','DG','DH','DI','DJ','DK','DL','DM','DN','DO','DP','DQ','DR','DS','DT','DU','DV','DW','DX','DY','DZ']
    end=2
    for j in range(0,len(ROI_list)):
            new_end=1   #standard set:min parameter length possible
            i=0
            k=0
            ws[abc[i]+'1']='dataset origin'
            for k in range(0,len(peak_time[j])):
                ws[abc[i]+str(k+end)]=name
            i+=1
            k=0
            ws[abc[i]+'1']='roi_origin'
            for k in range(0,len(peak_time[j])):
                ws[abc[i]+str(k+end)]=ROI_list[j]
            i+=1
            k=0
            ws[abc[i]+'1']='peak_time'
            for k in range(0,len(peak_time[j])):
                ws[abc[i]+str(k+end)]=peak_time[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='peak_value'
            for k in range(0,len(peak_value[j])):
                ws[abc[i]+str(k+end)]=peak_value[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='peak_begin_time'
            for k in range(0,len(peak_begin_time[j])):
                ws[abc[i]+str(k+end)]=peak_begin_time[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='peak_begin_value'
            for k in range(0,len(peak_begin_value[j])):
                ws[abc[i]+str(k+end)]=peak_begin_value[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='peak_end_time'
            for k in range(0,len(peak_end_time[j])):
                ws[abc[i]+str(k+end)]=peak_end_time[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='peak_end_value'
            for k in range(0,len(peak_end_value[j])):
                ws[abc[i]+str(k+end)]=peak_end_value[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='amplitude'
            for k in range(0,len(amplitude[j])):
                ws[abc[i]+str(k+end)]=amplitude[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='FWHM'
            for k in range(0,len(fwhm[j])):
                if fwhm[j][k]:
                    ws[abc[i]+str(k+end)]=fwhm[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='full_width'
            for k in range(0,len(full_width[j])):
                ws[abc[i]+str(k+end)]=full_width[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='area_under_curve'
            for k in range(0,len(auc[j])):
                ws[abc[i]+str(k+end)]=auc[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='risetime'
            for k in range(0,len(risetime[j])):
                ws[abc[i]+str(k+end)]=risetime[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='decay'
            for k in range(0,len(decay[j])):
                ws[abc[i]+str(k+end)]=decay[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='rise_20_time'
            for k in range(0,len(rise20_time[j])):
                ws[abc[i]+str(k+end)]=rise20_time[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='rise_20_value'
            for k in range(0,len(rise20_value[j])):
                ws[abc[i]+str(k+end)]=rise20_value[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='rise_80_time'
            for k in range(0,len(rise80_time[j])):
                ws[abc[i]+str(k+end)]=rise80_time[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='rise_80_value'
            for k in range(0,len(rise80_value[j])):
                ws[abc[i]+str(k+end)]=rise80_value[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='risetime2080'
            for k in range(0,len(risetime2080[j])):
                ws[abc[i]+str(k+end)]=risetime2080[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='fall_20_time'
            for k in range(0,len(fall20_time[j])):
                ws[abc[i]+str(k+end)]=fall20_time[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='fall_20_value'
            for k in range(0,len(fall20_value[j])):
                ws[abc[i]+str(k+end)]=fall20_value[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='fall_80_time'
            for k in range(0,len(fall80_time[j])):
                ws[abc[i]+str(k+end)]=fall80_time[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='fall_80_value'
            for k in range(0,len(fall80_value[j])):
                ws[abc[i]+str(k+end)]=fall80_value[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='decay2080'
            for k in range(0,len(decay2080[j])):
                ws[abc[i]+str(k+end)]=decay2080[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='tau_approx'
            for k in range(0,len(tau_approx[j])):
                ws[abc[i]+str(k+end)]=tau_approx[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='tau_fit'
            for k in range(0,len(tau_fit[j])):
                ws[abc[i]+str(k+end)]=tau_fit[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='responsetime'
            for k in range(0,len(responsetime[j])):
                ws[abc[i]+str(k+end)]=responsetime[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='time between peaks'
            for k in range(0,len(time_between_peaks[j])):
                ws[abc[i]+str(k+end)]=time_between_peaks[j][k]
            
            i+=1
            k=0
            ws[abc[i]+'1']='mean ISI'
            if mean_isi[j]:
                ws[abc[i]+str(end)]=mean_isi[j]
            i+=1
            k=0
            ws[abc[i]+'1']='median ISI'
            if median_isi[j]:
                ws[abc[i]+str(end)]=median_isi[j]
            i+=1
            k=0
            ws[abc[i]+'1']='ISI SD'
            if sd_isi[j]:
                ws[abc[i]+str(end)]=sd_isi[j]
            i+=1
            k=0
            ws[abc[i]+'1']='mean frequency'
            if mean_freq[j]:
                ws[abc[i]+str(end)]=mean_freq[j]
            i+=1
            k=0
            ws[abc[i]+'1']='median frequency'
            if median_freq[j]:
                ws[abc[i]+str(end)]=median_freq[j]
            i+=1
            k=0
            ws[abc[i]+'1']='frequency SD'
            if sd_freq[j]:
                ws[abc[i]+str(end)]=sd_freq[j]
            i+=1
            k=0
            ws[abc[i]+'1']='periodicity'
            if periodicity[j]:
                ws[abc[i]+str(end)]=periodicity[j]
            i+=1
            param_lengths=[len(peak_index[j])]
            new_end=max(param_lengths)
            end+=new_end
    wb.save(output_folder+'/'+name+'_results.xlsx')
    
    # # # # # # # # # # # # # # # # # # # # # # #
    # # # # #    save deleted peaks     # # # # #
    # # # # # # # # # # # # # # # # # # # # # # #

    wb=Workbook()
    ws=wb.active
    ws['A1']='dataset_origin'            
    ws['B1']='ROI origin'
    ws['C1']='deleted peak indices'
    counter=2
    for j in range(0,len(ROI_list)):
        for k in range(0,len(deleted_peaks[j])):
            ws['A'+str(counter)]=name
            ws['B'+str(counter)]=ROI_list[j]
            ws['C'+str(counter)]=deleted_peaks[j][k]
            counter+=1
    wb.save(output_folder+'/'+'peak_log_'+name+'.xlsx')

    
    # # # # # # # # # # # # # # # # # # # # # # # #
    # # # # #    save deleted columns     # # # # #
    # # # # # # # # # # # # # # # # # # # # # # # #

    wb=Workbook()
    ws=wb.active
    ws['A1']='dataset_origin'
    ws['B1']='deleted columns'
    counter=2
    for j in range(0,len(deleted_columns)):
        ws['A'+str(counter)]=name
        ws['B'+str(counter)]=deleted_columns[j]
        counter+=1
    wb.save(output_folder+'/'+'column_log_'+name+'.xlsx')

# # # # # # # # # # # # # # # # # # # # # # # #
# # # # # Explanation Analysis Skript # # # # #
# # # # # # # # # # # # # # # # # # # # # # # #

# 1. Input / Output [lines 421-434]
# Get path from user for the folders to read from and save in.

# 2. Do the following for all files in the input folder:

# 3. read in file [line 500-505, 22-53]

# 4. calculate time axis [lines 508-525]
# check if data of first column is linear -> time column and calculate framerate. Prompt if the framerate is incorrect or the user wants to use their own.

# 5. Do the following for all columns (except the time column) of the current file:

# 6. Data curation [lines 535-545]
# Display data and ask if the user wants to keep or discard the column. If discarded, add the ROI name to the deleted column memory. If not, continue with the analysis.

# 7. Correct the baseline shift [lines 599, 60-64]
# Use the function imodpoly() from the package pybaselines (authors) to calculate a baseline to subtract from the data.

# 8. Threshold definition [lines 606-613]
# To not calculate mean and sd from the whole trace, take all datapoints within the vertical border-values entered by the user and calculate mean and sd new from these selected datapoints.

# 9. Peak detection [lines 615-616, 65-75, 347-383]
# Find total maxima within data snippets that lie above the threshold of mean+3*sd

# 10. Peak curation [lines 623-652]
# Ask the user if these peaks should be kept as is or manually edited. If the user chooses so, they enter a range of two coordinates. All peaks within the rectangle build with these two diagonal coordinates will be deleted.

# 11. Peak edge detection [lines 655-657, 78-200, 404-409]
# Find the first point below the mean for each peak. If the trace between two peaks never crosses this threshold, approximate a peak beginning and end by finding the local minimum there.

# 12. Parameter calculation [lines 665-689, 202-412]
# Amplitude: 
# distance from peak value to average of peak begin and end value.
# full width at half max:
# Calculate crossings of half amplitude with the trace, then subtract right from left border.
# full width:
# Time of peak end subtracted from time of peak begin.
# area under curve:
# Area under curve is calculated for the area from peak begin to peak end, using an integration according to composite Simpson’s rule. It is calculated using the simpson() function from the package scipy.integrate.
# risetime:
# Time of peak subtracted from time of peak begin.
# decay:
# Time of peak end subtracted from time of peak.
# Risetime 20-80%:
# Detection of first datapoints that cross the thresholds at 20% and 80% amplitude. Subtraction of 80% from 20% timepoint.
# Decay 20-80%:
# Detection of last datapoints that cross the thresholds at 20% and 80% amplitude. Subtraction of 20% from 80% timepoint.
# Approximated tau:
# Calculate the threshold value by subtracting amplitude*0,632 from the peak value. Then find crossings with the decay area and calculate the time between peak and the crossing point.
# Tau calculated from fit:
# Fit an exponential function to the decay using the function curve_fit() from the package scipy.optimize and take tau from the returned parameters.
# Time between peaks:
# Subtract the previous peak time from the next for all detected peaks.
# Mean, median and sd of inter-event-interval:
# Calculate mean, median and sd of the time between peaks.
# Mean, median and sd of frequency:
# Calculate 1/mean, 1/median and 1/sd of the time between peaks.
# Periodicity:
# Periodicity is zero if sd of inter-event-interval is zero. Otherwise it is calculated by dividing sd of inter-event-interval by the mean of the inter-event-interval

# 13. Create results [lines 697-943]
# Write all parameters into the result excel sheet. Write all deleted peaks into the peak_log excel sheet. Write all deleted columns into the column_log excel sheet.
